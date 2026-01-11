"""Excel XLSX backend implementation for FinancesPy."""

import pathlib
import shutil
from collections.abc import Generator, Iterator
from contextlib import contextmanager
from datetime import date as datetime_date
from datetime import datetime
from typing import Any, Optional, Union

from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.worksheet import Worksheet

from financespy.backend import Backend
from financespy.categories import Categories, Category
from financespy.exceptions import (
    BackendConnectionError,
    BackendError,
    DataValidationError,
)
from financespy.money import Money
from financespy.transaction import Transaction


class XLSXBackend(Backend):
    """Excel XLSX-based storage backend for transactions.

    Stores transactions in Excel workbooks organized by year:
    - root_folder/YYYY.xlsx (with sheets: jan, feb, mar, etc.)

    Each sheet contains transactions for that month with columns:
    - Day, Categories, Value, Description
    """

    # Month names for sheet organization
    MONTH_NAMES = [
        "jan",
        "feb",
        "mar",
        "apr",
        "may",
        "jun",
        "jul",
        "aug",
        "sep",
        "oct",
        "nov",
        "dec",
    ]

    def __init__(
        self, folder: Union[str, pathlib.Path], categories: Categories
    ) -> None:
        """Initialize XLSX backend.

        Args:
            folder: Root folder path for storing Excel files
            categories: Category system for transaction categorization

        Raises:
            DataValidationError: If folder path is invalid
            BackendConnectionError: If folder cannot be accessed or created
        """
        if not folder:
            raise DataValidationError("Folder path cannot be empty")

        super().__init__(categories)

        self.folder = pathlib.Path(folder)
        self.categories = categories
        self._workbooks: dict[int, Workbook] = {}

        try:
            self.folder.mkdir(parents=True, exist_ok=True)

            test_file = self.folder / ".write_test"
            try:
                test_file.touch()
                test_file.unlink()
            except OSError as e:
                raise BackendConnectionError(
                    f"Cannot write to folder {folder}: {e}"
                ) from e

        except Exception as e:
            self._logger.error(f"Failed to initialize XLSX backend: {e}")
            raise BackendConnectionError(
                f"XLSX backend initialization failed: {e}"
            ) from e

        self._logger.info(f"XLSXBackend initialized with folder: {self.folder}")

    def insert_record(
        self, date: datetime_date, transaction: Transaction
    ) -> Optional[str]:
        """Insert a transaction record to the appropriate Excel file.

        Args:
            date: Transaction date
            transaction: Transaction to insert

        Returns:
            Generated transaction ID based on date and row position

        Raises:
            DataValidationError: If input validation fails
            BackendError: If file operation fails
        """
        if not isinstance(date, datetime_date):
            raise DataValidationError(f"Date must be a date object, got {type(date)}")

        if not isinstance(transaction, Transaction):
            raise DataValidationError(
                f"Transaction must be a Transaction object, got {type(transaction)}"
            )

        try:
            workbook = self._get_workbook(date)
            sheet = self._get_or_create_sheet(workbook, date.month)

            # Add transaction data to sheet
            sheet.append(
                [
                    date.day,
                    str(transaction.main_category()) if transaction.categories else "",
                    float(transaction.value),
                    transaction.description or "",
                ]
            )

            # Sort the sheet by day for better organization
            self._sort_sheet_by_day(sheet)

            # Save the workbook
            self._save_workbook(date.year, workbook)

            # Generate ID based on date and position
            transaction_id = f"{date.isoformat()}_{self._get_last_row_index(sheet)}"

            self._logger.debug(f"Inserted transaction to {self._get_filename(date)}")
            return transaction_id

        except Exception as e:
            self._logger.error(f"Failed to insert record for date {date}: {e}")
            raise BackendError(f"Insert operation failed: {e}") from e

    def records(self, date: datetime_date) -> Iterator[Transaction]:
        """Get all records for a specific date.

        Args:
            date: Date to query

        Returns:
            Iterator of transactions for the date

        Raises:
            DataValidationError: If date is invalid
            BackendError: If file operation fails
        """
        if not isinstance(date, datetime_date):
            raise DataValidationError(f"Date must be a date object, got {type(date)}")

        try:
            filename = self._get_filename(date)
            if not filename.exists():
                self._logger.debug(f"No Excel file found for date {date}")
                return iter([])

            workbook = self._get_workbook(date)
            sheet = workbook.worksheets[date.month - 1]

            return self._read_transactions_from_sheet(sheet, date)

        except Exception as e:
            self._logger.error(f"Failed to retrieve records for date {date}: {e}")
            raise BackendError(f"Query operation failed: {e}") from e

    def edit_record(self, transaction: Transaction) -> None:
        """Edit an existing transaction record.

        Args:
            transaction: Transaction with updated fields (must have valid ID)

        Raises:
            DataValidationError: If transaction validation fails
            BackendError: If edit operation fails
        """
        if not isinstance(transaction, Transaction):
            raise DataValidationError("Transaction must be a Transaction instance")

        if not hasattr(transaction, "id") or not transaction.id:
            raise DataValidationError("Transaction must have a valid ID")

        try:
            date_str, row_str = transaction.id.split("_")
            transaction_date = datetime.strptime(date_str, "%Y-%m-%d").date()
            row_index = int(row_str)

            if hasattr(transaction, "date") and transaction.date != transaction_date:
                self._edit_record_with_new_date(transaction)
                return

            workbook = self._get_workbook(transaction_date)
            sheet = workbook.worksheets[transaction_date.month - 1]

            # Update fields if provided
            if transaction.categories:
                sheet.cell(row=row_index, column=2).value = ", ".join(
                    str(cat) for cat in transaction.categories
                )

            if transaction.value:
                sheet.cell(row=row_index, column=3).value = float(transaction.value)

            if transaction.description is not None:
                sheet.cell(row=row_index, column=4).value = transaction.description

            self._save_workbook(transaction_date.year, workbook)
            self._logger.debug(f"Edited transaction {transaction.id}")

        except Exception as e:
            self._logger.error(f"Failed to edit transaction {transaction.id}: {e}")
            raise BackendError(f"Edit operation failed: {e}") from e

    def delete_record(self, transaction_id: str) -> None:
        """Delete a transaction record.

        Args:
            transaction_id: ID of the transaction to delete

        Raises:
            DataValidationError: If transaction_id is invalid
            BackendError: If delete operation fails
        """
        if not isinstance(transaction_id, str) or not transaction_id.strip():
            raise DataValidationError("Transaction ID must be a non-empty string")

        try:
            date_str, row_str = transaction_id.split("_")
            transaction_date = datetime.strptime(date_str, "%Y-%m-%d").date()
            row_index = int(row_str)

            workbook = self._get_workbook(transaction_date)
            sheet = workbook.worksheets[transaction_date.month - 1]

            sheet.delete_rows(row_index)
            self._save_workbook(transaction_date.year, workbook)

            self._logger.debug(f"Deleted transaction {transaction_id}")

        except Exception as e:
            self._logger.error(f"Failed to delete transaction {transaction_id}: {e}")
            raise BackendError(f"Delete operation failed: {e}") from e

    def _get_filename(self, date: datetime_date) -> pathlib.Path:
        """Get the Excel filename for a given date.

        Args:
            date: Date to get filename for

        Returns:
            Path to the Excel file
        """
        return self.folder / f"{date.year}.xlsx"

    def _get_template_path(self) -> pathlib.Path:
        """Get the path to the XLSX template file.

        Returns:
            Path to the template file
        """
        return pathlib.Path(__file__).parent / "template.xlsx"

    def _get_workbook(self, date: datetime_date) -> Workbook:
        """Get or load the workbook for a given date.

        Args:
            date: Date to get workbook for

        Returns:
            Excel Workbook object

        Raises:
            BackendError: If workbook cannot be loaded or created
        """
        if date.year not in self._workbooks:
            filename = self._get_filename(date)

            try:
                if filename.exists():
                    workbook = load_workbook(filename=filename)
                else:
                    # Copy from template if it exists, otherwise create new
                    template_path = self._get_template_path()
                    if template_path.exists():
                        shutil.copy(template_path, filename)
                        workbook = load_workbook(filename=filename)
                        self._logger.info(f"Created {filename} from template")
                    else:
                        workbook = self._create_new_workbook()
                        self._logger.warning(
                            f"Template not found at {template_path}, "
                            "creating blank workbook"
                        )

                self._workbooks[date.year] = workbook
                return workbook
            except InvalidFileException as e:
                raise BackendError(f"Invalid Excel file {filename}: {e}") from e
            except Exception as e:
                raise BackendError(f"Failed to load workbook {filename}: {e}") from e

        return self._workbooks[date.year]

    def _create_new_workbook(self) -> Workbook:
        """Create a new workbook with monthly sheets.

        Returns:
            New Workbook with 12 monthly sheets
        """
        workbook = Workbook()

        # Remove default sheet
        workbook.remove(workbook.active)

        for month_name in self.MONTH_NAMES:
            sheet = workbook.create_sheet(title=month_name)
            # Add header row
            sheet.append(["Day", "Categories", "Value", "Description"])

        return workbook

    def _get_or_create_sheet(self, workbook: Workbook, month: int) -> Worksheet:
        """Get or create a sheet for the given month.

        Args:
            workbook: Excel workbook
            month: Month number (1-12)

        Returns:
            Worksheet for the month

        Raises:
            DataValidationError: If month is invalid
        """
        if not 1 <= month <= 12:
            raise DataValidationError(f"Invalid month: {month}")

        try:
            return workbook.worksheets[month - 1]
        except IndexError:
            while len(workbook.worksheets) < month:
                sheet_name = self.MONTH_NAMES[len(workbook.worksheets)]
                sheet = workbook.create_sheet(title=sheet_name)
                sheet.append(["Day", "Categories", "Value", "Description"])

            return workbook.worksheets[month - 1]

    def _read_transactions_from_sheet(
        self, sheet: Worksheet, date: datetime_date
    ) -> Iterator[Transaction]:
        """Read transactions from an Excel sheet for a specific date.

        Args:
            sheet: Excel worksheet
            date: Date to filter transactions for

        Yields:
            Transaction objects for the specified date
        """
        if not self.categories:
            raise BackendError("Categories system required for reading transactions")

        try:
            rows = list(sheet.rows)
            if len(rows) <= 1:  # Only header or empty
                return

            for row_index, row in enumerate(rows[1:], start=2):  # Skip header
                transaction = self._parse_transaction_row(row, date, row_index)
                if transaction:
                    yield transaction

        except Exception as e:
            raise BackendError(f"Failed to read from sheet: {e}") from e

    def _parse_transaction_row(
        self, row: tuple[Any, ...], date: datetime_date, row_index: int
    ) -> Optional[Transaction]:
        """Parse a single row into a transaction if it matches the date."""
        try:
            day_cell, categories_cell, value_cell, description_cell = row[:4]

            if not day_cell.value or int(day_cell.value) != date.day:
                return None

            categories = self._parse_categories(categories_cell)
            value = Money(value_cell.value) if value_cell.value else Money(0)
            description = str(description_cell.value) if description_cell.value else ""

            transaction = Transaction(
                value=value, categories=categories, description=description
            )
            transaction.id = f"{date.isoformat()}_{row_index}"
            transaction.date = date

            return transaction

        except Exception as e:
            self._logger.warning(f"Failed to parse row {row_index} in sheet: {e}")
            return None

    def _parse_categories(self, categories_cell: Any) -> list[Category]:
        """Parse categories from a cell value."""
        categories = []
        if categories_cell.value:
            category_names = [
                name.strip() for name in str(categories_cell.value).split(",")
            ]
            categories = [
                self.categories.category(name) for name in category_names if name
            ]
        return categories

    def _sort_sheet_by_day(self, sheet: Worksheet) -> None:
        """Sort sheet rows by day column.

        Args:
            sheet: Excel worksheet to sort
        """
        try:
            data_rows = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if any(cell is not None for cell in row):  # Skip empty rows
                    data_rows.append(list(row))

            # Sort by day (first column)
            data_rows.sort(key=lambda row: row[0] if row[0] is not None else 0)

            # Clear sheet content (except header)
            sheet.delete_rows(2, sheet.max_row)

            # Rewrite sorted data
            for row_data in data_rows:
                sheet.append(row_data)

        except Exception as e:
            self._logger.warning(f"Failed to sort sheet: {e}")

    def _save_workbook(self, year: int, workbook: Workbook) -> None:
        """Save workbook to file.

        Args:
            year: Year for the workbook
            workbook: Workbook to save

        Raises:
            BackendError: If save operation fails
        """
        try:
            filename = self.folder / f"{year}.xlsx"
            workbook.save(filename)
        except Exception as e:
            raise BackendError(f"Failed to save workbook for year {year}: {e}") from e

    def _get_last_row_index(self, sheet: Worksheet) -> int:
        """Get the index of the last row with data.

        Args:
            sheet: Excel worksheet

        Returns:
            Row index of last data row
        """
        return int(sheet.max_row)

    def _edit_record_with_new_date(self, transaction: Transaction) -> None:
        """Handle editing a transaction with a new date.

        This involves deleting the old record and inserting a new one.

        Args:
            transaction: Transaction with new date
        """
        if not transaction.id:
            raise DataValidationError("Transaction ID is required")
        if not transaction.date:
            raise DataValidationError("Transaction date is required")

        old_id = transaction.id
        self.delete_record(old_id)
        new_id = self.insert_record(transaction.date, transaction)
        transaction.id = new_id

    def get_available_years(self) -> list[int]:
        """Get list of years that have transaction data.

        Returns:
            List of years with available Excel files
        """
        years = []
        try:
            for file_path in self.folder.glob("*.xlsx"):
                try:
                    year = int(file_path.stem)
                    years.append(year)
                except ValueError:
                    continue  # Skip non-year files
            return sorted(years)
        except Exception as e:
            self._logger.error(f"Failed to get available years: {e}")
            raise BackendError(f"Failed to get available years: {e}") from e

    @contextmanager
    def _atomic_workbook_operation(
        self, date: datetime_date
    ) -> Generator[Workbook, None, None]:
        """Context manager for atomic workbook operations.

        Ensures workbook is properly saved even if operation fails.

        Args:
            date: Date for workbook selection

        Yields:
            Workbook object
        """
        workbook = self._get_workbook(date)
        try:
            yield workbook
        finally:
            try:
                self._save_workbook(date.year, workbook)
            except Exception as e:
                self._logger.error(f"Failed to save workbook after operation: {e}")

    def close_all_workbooks(self) -> None:
        """Close all cached workbooks and clear cache.

        This helps free memory and ensures all changes are saved.
        """
        for year, workbook in self._workbooks.items():
            try:
                self._save_workbook(year, workbook)
            except Exception as e:
                self._logger.error(f"Failed to save workbook for year {year}: {e}")

        self._workbooks.clear()
        self._logger.info("Closed all workbooks")

    def __exit__(
        self,
        exc_type: Optional[type[BaseException]],
        exc_value: Optional[BaseException],
        traceback: Any,
    ) -> None:
        """Context manager exit.

        Args:
            exc_type: Exception type
            exc_value: Exception value
            traceback: Exception traceback
        """
        self.close_all_workbooks()
