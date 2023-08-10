from openpyxl import load_workbook

from financespy.backend import Backend
import pathlib
from financespy.transaction import Transaction, parse_transaction


def with_index(trans, index):
    trans.id = index
    return trans


class XLSXBackend(Backend):
    def __init__(self, folder):
        super().__init__()
        self.folder = pathlib.Path(folder)
        self._workbooks = {}

    def _filename(self, date):
        return self.folder / f"{str(date.year)}.xlsx"

    def _get_workbook(self, date):
        if date.year not in self._workbooks:
            workbook = load_workbook(filename=self._filename(date))
            self._workbooks[date.year] = workbook
            return workbook

        return self._workbooks[date.year]

    def _rows_to_records(self, rows, date):
        return (
            with_index(
                parse_transaction(
                    str(row[2].value) + "," + str(row[1].value), self.categories
                ),
                index,
            )
            for row, index in list(rows)[1:]
            if row[0].value and date.day == int(row[0].value)
        )

    def records(self, date):
        workbook = self._get_workbook(date)
        rows = list(workbook.worksheets[date.month - 1].rows)
        rows = zip(rows, range(0, len(rows)))
        return self._rows_to_records(rows, date)

    def insert_record(self, date, transaction):
        if type(transaction) is not Transaction:
            raise TypeError("Supplied parameter is not a transaction")

        workbook = self._get_workbook(date)
        sheet = workbook.worksheets[date.month - 1]
        sheet.append(
            [
                date.day,
                str(transaction.main_category()),
                str(transaction.value),
                transaction.description,
            ]
        )
        sheet_copy = [[c.value for c in r] for r in sheet][1:]
        sheet_copy.sort(key=lambda row: row[0])

        for i in range(0, len(sheet_copy)):
            for j in range(0, len(sheet_copy[i])):
                sheet.cell(i + 2, j + 1).value = sheet_copy[i][j]

        workbook.save(self._filename(date))

    def update_record(self, transaction):
        date = transaction.date
        workbook = self._get_workbook(date)
        sheet = workbook.worksheets[date.month - 1]

        sheet[transaction.id] = [
            date.day,
            str(transaction.main_category()),
            str(transaction.value),
            transaction.description,
        ]
